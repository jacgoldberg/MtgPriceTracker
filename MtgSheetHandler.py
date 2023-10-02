import requests
import openpyxl
from bs4 import BeautifulSoup

class CardSheet:
    def __init__(self, excelName):
        self.wb_obj = openpyxl.load_workbook(excelName)
        self.sheet_obj = self.wb_obj.active
    def update(self):
        for row in self.sheet_obj.iter_rows():
            if (row[0].value != None) & (row[0].value != "Set Name"):
                setName = row[0].value.replace(" ", "+")
            if (row[1].value != None) & (row[1].value != "Card Name"):
                cardName = row[1].value.replace(" ", "+").replace("'", "").replace(",", "")
            if (row[2].value != None) & (row[2].value != "Foil"):
                if(row[2].value):
                    foil = ":Foil"
                else:
                    foil = ""
            if (row[3].value != None) & (row[3].value != "Borderless"):
                if(row[3].value):
                    borderless = "-borderless"
                else:
                    borderless = "" 
            if (row[4].value != None) & (row[4].value != "Showcase"): 
                if(row[4].value):
                    showCase = "-showcase"
                else:
                    showCase = "" 
            if (row[5].value != "MTGGoldfishLink") & (row[0].value != None) & (row[1].value != None) & (row[2].value != None) & (row[3].value != None) & (row[4].value != None):
                URL = "https://www.mtggoldfish.com/price/{}{}/{}{}{}#paper".format(setName.strip("\n"),foil,cardName.strip("\n"),borderless,showCase)
                row[5].value = URL
                print(URL)
                page = requests.get(URL)
                soup = BeautifulSoup(page.content, "html.parser")
                results = soup.find("div", {"class": "price-box-price"})
                price = results.string
                row[7].value = float(price[2:7])
                if(row[6].value == None):
                    row[6].value = float(price[2:7])
                row[8].value = row[7].value - row[6].value
        self.wb_obj.save("MagicCardTracker.xlsx")
    def addCard(self, setName, cardName, foil, borderless, showCase):
        if foil == 1:
            foil = True
        else:
            foil = False
        if borderless == 1:
            borderless = True
        else:
            borderless = False
        if showCase == 1:
            showCase = True
        else:
            showCase = False
        for row in self.sheet_obj.iter_rows():
            if(row[0].value == None):
                row[0].value = setName
                row[1].value = cardName
                row[2].value = foil
                row[3].value = borderless
                row[4].value = showCase
                self.wb_obj.save("MagicCardTracker.xlsx")
                return
        self.sheet_obj.append({'A' : setName, 'B' : cardName, 'C' : foil, 'D' : borderless, 'E' : showCase})
        self.wb_obj.save("MagicCardTracker.xlsx")
                

    